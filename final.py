import streamlit as st
import os
from dotenv import load_dotenv
import base64
import json
import time
from mistralai import Mistral
from openpyxl import load_workbook
import pandas as pd 
st.set_page_config(page_title="📊 Excel RAG Chatbot", layout="wide")
st.title("BillBot")
def get_base64(bin_file):
    with open(bin_file, 'rb') as f:
        data = f.read()
    return base64.b64encode(data).decode()

def set_background(png_file):
    bin_str = get_base64(png_file)
    page_bg_img = '''
    <style>
    .stApp {
    background-color : #B0D8F3;
    background-image: url("data:image/png;base64,%s");
    background-size: cover;
    }
    </style>
    ''' % bin_str
    st.markdown(page_bg_img, unsafe_allow_html=True)


set_background('D:\\quantumX2025\\background\\Slide2.PNG')
# 1. API Key Input
api_key = "AUHGTFQUOgInIPfa1PIgH0F6UaKsGGz3"

# Initialize session state variables for persistence
if "ocr_result" not in st.session_state:
    st.session_state["ocr_result"] = []
if "preview_src" not in st.session_state:
    st.session_state["preview_src"] = []
if "image_bytes" not in st.session_state:
    st.session_state["image_bytes"] = []

# 2. Choose file type: PDF or Image
file_type = st.radio("Select file type", ("PDF", "Image"))

# 3. Select source type: URL or Local Upload
source_type = st.radio("Select source type", ("URL", "Local Upload"))

input_url = ""
uploaded_files = []

if source_type == "URL":
    input_url = st.text_area("Enter one or multiple URLs (separate with new lines)")
else:
    uploaded_files = st.file_uploader("Upload one or more files", type=["pdf", "jpg", "jpeg", "png"], accept_multiple_files=True)

# Generate CSV output 



# 4. Process Button & OCR Handling
if st.button("Process"):
    if source_type == "URL" and not input_url.strip():
        st.error("Please enter at least one valid URL.")
    elif source_type == "Local Upload" and not uploaded_files:
        st.error("Please upload at least one file.")
    else:
        #MISTRAL AI IS HERE
        client = Mistral(api_key=api_key)
        st.session_state["ocr_result"] = []
        st.session_state["preview_src"] = []
        st.session_state["image_bytes"] = []
        
        sources = input_url.split("\n") if source_type == "URL" else uploaded_files
        
        for idx, source in enumerate(sources):
            if file_type == "PDF":
                if source_type == "URL":
                    document = {"type": "document_url", "document_url": source.strip()}
                    preview_src = source.strip()
                else:
                    file_bytes = source.read()
                    encoded_pdf = base64.b64encode(file_bytes).decode("utf-8")
                    document = {"type": "document_url", "document_url": f"data:application/pdf;base64,{encoded_pdf}"}
                    preview_src = f"data:application/pdf;base64,{encoded_pdf}"
            else:
                if source_type == "URL":
                    document = {"type": "image_url", "image_url": source.strip()}
                    preview_src = source.strip()
                else:
                    file_bytes = source.read()
                    mime_type = source.type
                    encoded_image = base64.b64encode(file_bytes).decode("utf-8")
                    document = {"type": "image_url", "image_url": f"data:{mime_type};base64,{encoded_image}"}
                    preview_src = f"data:{mime_type};base64,{encoded_image}"
                    st.session_state["image_bytes"].append(file_bytes)
            
            with st.spinner(f"Processing {source if source_type == 'URL' else source.name}..."):
                try:
                    #CHANGE TO PADDLEOCR 
                    ocr_response = client.ocr.process(model="mistral-ocr-latest", document=document, include_image_base64=True )
                    time.sleep(1)  # wait 1 second between request to prevent rate limit exceeding
                    
                    pages = ocr_response.pages if hasattr(ocr_response, "pages") else (ocr_response if isinstance(ocr_response, list) else [])
                    result_text = "\n\n".join(page.markdown for page in pages) or "No result found."
                except Exception as e:
                    result_text = f"Error extracting result: {e}"
                
                st.session_state["ocr_result"].append(result_text)
                st.session_state["preview_src"].append(preview_src)
                #print("output: \n",result_text)
                #print(type(result_text))
        
                # providing prompt to mistral ai 

                prompt = f"""
                            You are a bills and invoice parser. Understand raw and unstructured text and extract structured fields from the following bill text. map the fields to the following JSON format.
                            if there are fields you can identify but not in the JSON format, add them to the JSON as well. add only relevant fields to the JSON.
                            If you cannot identify any fields, remove that JSON field from the answer. do not generate own content.
                            Return the output as JSON with fields like this and all the fields should be seperately mapped (no nested json). Stick to the order of the fields mentioned below.:
                            - vendor_name
                            - invoice_number
                            - date
                            - due_date
                            - total_amount
                            - description
                            - quantity
                            - unit_price
                            - tax
                            - subtotal
                            - currency

                            give output only in json format. do not add any other text.
                            Bill Text:
                            \"\"\"
                            {result_text}
                            \"\"\"
                            """

                model = "mistral-large-latest"
                client = Mistral(api_key=api_key)

                chat_response = client.chat.complete(
                    model = model,
                    messages = [
                        {
                            "role": "user",
                            "content": prompt,
                        },
                    ]
                )

                response = chat_response.choices[0].message.content
                import json
                with open('data.json', 'w') as f:
                    json.dump(response, f)
                #print(response)
                
                lines = response.strip().splitlines()
                # Remove the first and last training lines
                clean_json_str = "\n".join(lines[1:-1])

                # Parse the cleaned JSON string
                response = json.loads(clean_json_str)
                print(response)
               # 
                # Assuming `df` is the DataFrame you want to append
                df = pd.json_normalize(response)

                # Define the path to your existing Excel file
                file_path = 'D:\\quantumX2025\\inv_small.xlsx'
                

                
                try:
                    # Open the existing Excel file using openpyxl
                    book = load_workbook(file_path)

                    # Check if the sheet exists, if it does, append to it
                    if "Sheet1" in book.sheetnames:
                        # Get the existing sheet
                        sheet = book["Sheet1"]
                        
                        # Start appending data after the last row
                        start_row = sheet.max_row

                        # Open ExcelWriter in append mode
                        df.columns = df.columns.map(str)
                        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
                            # Append data to the existing sheet, starting after the last row
                            df.to_excel(writer, sheet_name='Sheet1', startrow=start_row, index=False, header=False)
                            writer.save()

                    else:
                        # If "Sheet1" doesn't exist, create a new sheet
                        df.columns = df.columns.map(str)
                        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
                            df.to_excel(writer, sheet_name='Sheet1', startrow=start_row, index=False, header=False)
                            writer.save()

                    print("Data appended successfully!")
                except PermissionError:
                    print("PermissionError: Please ensure the file is not open in Excel.")
                except Exception as e:
                    print(f"Error: {e}")
                
import streamlit as st
import pandas as pd
import numpy as np
import requests
import faiss
import json

from langchain import hub
from langchain_core.output_parsers import StrOutputParser
from langchain_core.runnables import RunnablePassthrough
from langchain_core.prompts import ChatPromptTemplate
from langchain_ollama import ChatOllama


model = ChatOllama(model="deepseek-r1", base_url="http://localhost:11434")



prompt = hub.pull("rlm/rag-prompt")
prompt = '''You are a helpful data analyst.
        Given the following rows from an Excel file, answer the user's question using only this data.

        Data:
        {context}

        Question: {query}'''

prompt = ChatPromptTemplate.from_template(prompt)

def format_docs(docs):
    return "\n\n".join([doc.page_content for doc in docs])

    # print(format_docs(docs))

rag_chain = (
    prompt
    | model
    | StrOutputParser()
)


# === Local Nomic Embed
def get_embedding(texts):
    embeddings = []
    for text in texts:
        response = requests.post(
            "http://localhost:11434/api/embeddings",
            json={"model": "nomic-embed-text", "prompt": text}
        )
        response.raise_for_status()
        embeddings.append(response.json()["embedding"])
    return np.array(embeddings)

# === Streamlit UI

#st.title("📊 Excel RAG Chatbot (Local Nomic Embed + Local DeepSeek-Coder)")

#uploaded_file = st.file_uploader("Upload your Excel file", type=[".xlsx", ".xls"])

#if uploaded_file:
    


df = pd.read_excel('D:\\quantumX2025\\inv_small.xlsx')
df = df.astype(str)  # ensure all data is string for embedding
st.dataframe(df.head(10))

    # Convert rows to text
data_chunks = [row.to_json() for _, row in df.iterrows()]

with st.spinner("🔍 Embedding data with nomic-embed-text..."):
    embeddings = get_embedding(data_chunks)

        # Build FAISS index
    dimension = len(embeddings[0])
    index = faiss.IndexFlatL2(dimension)
    index.add(embeddings)

        # Input query
    query = st.text_input("Ask a question about your Excel data")
    if query:
        with st.spinner("🔍 Embedding query..."):
            query_embedding = get_embedding([query])[0]
            D, I = index.search(np.array([query_embedding]), k=5)

            # Get top matching chunks
            retrieved_chunks = [data_chunks[i] for i in I[0]]
            context = "\n".join(retrieved_chunks)

            # Build prompt
            prompt = f"""
            You are a helpful data analyst.
            Do not add thinking process in the final output. give the answer to the query only. keep it short and
            stick to maximum two lines (one line recommeneded). do not elaborate on the answer.
            Given the following rows from an Excel file, answer the user's question using only this data.

            Data:
            {context}

            Question: {query}
            """

            with st.spinner("💬 Generating response from DeepSeek (local)..."):
                answer = rag_chain.invoke({
                    "query": query,
                    "context": context
                })
                cleaned_answer = answer.replace("<|thinking|>", "").strip()
                st.markdown(f"**🧠 DeepSeek:** {cleaned_answer}")
